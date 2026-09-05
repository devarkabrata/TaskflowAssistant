"""Writes an LLM-decided table (columns + rows) out as an .xlsx workbook, via openpyxl.

`ExportDocumentInput`'s validator guarantees `columns`/`rows` are both present
whenever `format="excel"` — there is no fallback path here for freeform
`content`, a spreadsheet has to be tabular.
"""

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from taskflowassistant.dedicated_tools.table_utils import row_value

_MAX_SHEET_TITLE_LENGTH = 31  # Excel's own hard limit on sheet names.

# Excel rejects a sheet title containing any of : \ / ? * [ ] outright
# (openpyxl raises ValueError) — observed in production: the model's own
# document title (e.g. "Sprint 1: Backlog") often contains one of these,
# since nothing about the tool's schema tells it not to. The sheet title is
# purely internal/cosmetic (the user-facing filename comes from
# builder.py's own `_slugify(title)`), so stripping these here is safe and
# means a natural title no longer crashes the whole export.
_INVALID_SHEET_TITLE_CHARS = re.compile(r"[:\\/?*\[\]]")


def _sanitize_sheet_title(title: str) -> str:
    cleaned = _INVALID_SHEET_TITLE_CHARS.sub(" ", title).strip()
    return cleaned[:_MAX_SHEET_TITLE_LENGTH] or "Sheet1"


def write_excel_file(title: str, columns: list[str], rows: list[dict], path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _sanitize_sheet_title(title)

    sheet.append(columns)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append([row_value(row, column) for column in columns])

    for idx in range(1, len(columns) + 1):
        sheet.column_dimensions[get_column_letter(idx)].width = 24

    workbook.save(path)
    return path
